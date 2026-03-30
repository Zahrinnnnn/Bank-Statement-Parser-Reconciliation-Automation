"""
test_reports.py — Tests for the Excel and PDF report generators.

Run with:
    python -m pytest tests/test_reports.py -v
"""

import tempfile
from datetime import date, datetime
from pathlib import Path

import pytest

from src.database.connection import DatabaseConnection
from src.database.models import BankTransaction, LedgerEntry
from src.database.queries import (
    insert_bank_transaction,
    insert_ledger_entry,
    get_reconciliation,
    list_match_details_for_reconciliation,
    list_exceptions_for_reconciliation,
    list_bank_transactions,
)
from src.reconciliation.engine import run_reconciliation
from src.reports.excel_report import generate_excel_report
from src.reports.pdf_report import generate_pdf_report


# ---------------------------------------------------------------------------
# Shared fixture — in-memory DB with a completed reconciliation run
# ---------------------------------------------------------------------------

class ReportTestBase:
    """
    Base class that sets up an in-memory database with known transactions,
    runs reconciliation, and exposes everything report generators need.
    """

    def setup_method(self):
        self.db = DatabaseConnection(db_path=Path(":memory:"))
        self.db.connect()
        self.db.initialise_schema()
        self.conn = self.db.get_connection()
        self._load_sample_data()
        self._run_reconciliation()

    def teardown_method(self):
        self.db.close()

    def _load_sample_data(self):
        bank_txns = [
            BankTransaction(
                bank_name="CIMB", account_number="1234567890",
                transaction_date=date(2026, 3, 3),
                description="FPX SHOPEE PAYMENT", reference="FPX2026001",
                debit_amount=250.0, credit_amount=0.0,
                source_file="test.csv", hash="rpt_hash_shopee",
            ),
            BankTransaction(
                bank_name="CIMB", account_number="1234567890",
                transaction_date=date(2026, 3, 5),
                description="SALARY CREDIT ACME SDN BHD", reference=None,
                debit_amount=0.0, credit_amount=3500.0,
                source_file="test.csv", hash="rpt_hash_salary",
            ),
            BankTransaction(
                bank_name="CIMB", account_number="1234567890",
                transaction_date=date(2026, 3, 8),
                description="IBG TRANSFER OUT", reference="IBG20260308",
                debit_amount=1000.0, credit_amount=0.0,
                source_file="test.csv", hash="rpt_hash_ibg",
            ),
            # This one has no ledger match — will become BANK_ONLY exception
            BankTransaction(
                bank_name="CIMB", account_number="1234567890",
                transaction_date=date(2026, 3, 20),
                description="UNMATCHED DEBIT NO LEDGER", reference=None,
                debit_amount=99.0, credit_amount=0.0,
                source_file="test.csv", hash="rpt_hash_unknown",
            ),
        ]
        for txn in bank_txns:
            insert_bank_transaction(self.conn, txn)

        ledger_entries = [
            LedgerEntry(
                entry_date=date(2026, 3, 3), description="Shopee Payment March",
                reference="FPX2026001", amount=250.0, entry_type="DEBIT",
            ),
            LedgerEntry(
                entry_date=date(2026, 3, 5), description="Salary March 2026",
                reference="PAYROLL-MAR-26", amount=3500.0, entry_type="CREDIT",
            ),
            LedgerEntry(
                entry_date=date(2026, 3, 8), description="Transfer to Ahmad Ali",
                reference="IBG20260308", amount=1000.0, entry_type="DEBIT",
            ),
            # This one has no bank match — will become LEDGER_ONLY exception
            LedgerEntry(
                entry_date=date(2026, 3, 15),
                description="Outstanding vendor payment",
                reference=None, amount=500.0, entry_type="DEBIT",
            ),
        ]
        for entry in ledger_entries:
            insert_ledger_entry(self.conn, entry)

    def _run_reconciliation(self):
        self.result = run_reconciliation(
            conn=self.conn,
            period_start=date(2026, 3, 1),
            period_end=date(2026, 3, 31),
            bank_name="CIMB",
            account_number="1234567890",
        )
        self.recon_id    = self.result.recon_id
        self.recon       = dict(get_reconciliation(self.conn, self.recon_id))
        self.matched_rows    = list_match_details_for_reconciliation(self.conn, self.recon_id)
        self.exception_rows  = list_exceptions_for_reconciliation(self.conn, self.recon_id)
        self.all_bank_txns   = list_bank_transactions(
            self.conn,
            bank_name="CIMB",
            account_number="1234567890",
            period_start=date(2026, 3, 1),
            period_end=date(2026, 3, 31),
        )


# ===========================================================================
# list_match_details_for_reconciliation — queries.py
# ===========================================================================

class TestListMatchDetails(ReportTestBase):

    def test_returns_correct_number_of_matched_pairs(self):
        # Three pairs should be matched (Shopee, Salary, IBG)
        assert len(self.matched_rows) == 3

    def test_each_row_has_required_keys(self):
        required_keys = [
            "match_id", "match_type", "confidence_score",
            "bank_txn_id", "bank_date", "bank_description",
            "debit_amount", "credit_amount", "bank_reference",
            "ledger_entry_id", "ledger_date", "ledger_description",
            "ledger_amount", "ledger_entry_type", "ledger_reference",
        ]
        for row in self.matched_rows:
            for key in required_keys:
                assert key in row, f"Missing key '{key}' in match row"

    def test_shopee_row_has_correct_match_type(self):
        shopee = next(r for r in self.matched_rows if "SHOPEE" in r["bank_description"])
        assert shopee["match_type"] == "EXACT"
        assert shopee["confidence_score"] == 1.0

    def test_amounts_are_populated(self):
        for row in self.matched_rows:
            # Every matched row must have a non-negative amount on at least one side
            bank_total = (row["debit_amount"] or 0) + (row["credit_amount"] or 0)
            assert bank_total > 0

    def test_no_match_rows_for_unmatched_exception(self):
        # The UNMATCHED DEBIT row has no ledger match so should not appear here
        descriptions = [r["bank_description"] for r in self.matched_rows]
        assert not any("UNMATCHED" in d for d in descriptions)


# ===========================================================================
# Excel report — file output and content checks
# ===========================================================================

class TestExcelReport(ReportTestBase):

    def test_file_is_created(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            output_path = Path(tmp_dir) / "test_recon.xlsx"
            generate_excel_report(
                recon=self.recon,
                matched_rows=self.matched_rows,
                exception_rows=self.exception_rows,
                all_bank_txns=self.all_bank_txns,
                output_path=output_path,
            )
            assert output_path.exists()

    def test_file_is_non_empty(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            output_path = Path(tmp_dir) / "test_recon.xlsx"
            generate_excel_report(
                recon=self.recon,
                matched_rows=self.matched_rows,
                exception_rows=self.exception_rows,
                all_bank_txns=self.all_bank_txns,
                output_path=output_path,
            )
            assert output_path.stat().st_size > 0

    def test_returns_output_path(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            output_path = Path(tmp_dir) / "test_recon.xlsx"
            returned = generate_excel_report(
                recon=self.recon,
                matched_rows=self.matched_rows,
                exception_rows=self.exception_rows,
                all_bank_txns=self.all_bank_txns,
                output_path=output_path,
            )
            assert returned == output_path

    def test_all_four_sheets_exist(self):
        """Open the workbook with openpyxl and check sheet names."""
        import openpyxl

        with tempfile.TemporaryDirectory() as tmp_dir:
            output_path = Path(tmp_dir) / "test_recon.xlsx"
            generate_excel_report(
                recon=self.recon,
                matched_rows=self.matched_rows,
                exception_rows=self.exception_rows,
                all_bank_txns=self.all_bank_txns,
                output_path=output_path,
            )
            wb = openpyxl.load_workbook(output_path)
            sheet_names = wb.sheetnames
            assert "Summary" in sheet_names
            assert "Matched Transactions" in sheet_names
            assert "Exceptions" in sheet_names
            assert "All Transactions" in sheet_names

    def test_matched_sheet_has_correct_row_count(self):
        """Matched Transactions sheet should have 1 header + 3 data rows."""
        import openpyxl

        with tempfile.TemporaryDirectory() as tmp_dir:
            output_path = Path(tmp_dir) / "test_recon.xlsx"
            generate_excel_report(
                recon=self.recon,
                matched_rows=self.matched_rows,
                exception_rows=self.exception_rows,
                all_bank_txns=self.all_bank_txns,
                output_path=output_path,
            )
            wb   = openpyxl.load_workbook(output_path)
            ws   = wb["Matched Transactions"]
            # max_row includes the header row
            assert ws.max_row == 1 + len(self.matched_rows)

    def test_exceptions_sheet_has_correct_row_count(self):
        """Exceptions sheet should have 1 header + 2 exception rows."""
        import openpyxl

        with tempfile.TemporaryDirectory() as tmp_dir:
            output_path = Path(tmp_dir) / "test_recon.xlsx"
            generate_excel_report(
                recon=self.recon,
                matched_rows=self.matched_rows,
                exception_rows=self.exception_rows,
                all_bank_txns=self.all_bank_txns,
                output_path=output_path,
            )
            wb = openpyxl.load_workbook(output_path)
            ws = wb["Exceptions"]
            assert ws.max_row == 1 + len(self.exception_rows)

    def test_all_transactions_sheet_has_correct_row_count(self):
        """All Transactions sheet should have 1 header + 4 bank transaction rows."""
        import openpyxl

        with tempfile.TemporaryDirectory() as tmp_dir:
            output_path = Path(tmp_dir) / "test_recon.xlsx"
            generate_excel_report(
                recon=self.recon,
                matched_rows=self.matched_rows,
                exception_rows=self.exception_rows,
                all_bank_txns=self.all_bank_txns,
                output_path=output_path,
            )
            wb = openpyxl.load_workbook(output_path)
            ws = wb["All Transactions"]
            assert ws.max_row == 1 + len(self.all_bank_txns)

    def test_output_path_parent_created_if_missing(self):
        """generate_excel_report should create any missing parent directories."""
        with tempfile.TemporaryDirectory() as tmp_dir:
            nested_path = Path(tmp_dir) / "new_folder" / "sub" / "report.xlsx"
            generate_excel_report(
                recon=self.recon,
                matched_rows=self.matched_rows,
                exception_rows=self.exception_rows,
                all_bank_txns=self.all_bank_txns,
                output_path=nested_path,
            )
            assert nested_path.exists()

    def test_report_works_with_empty_matched_rows(self):
        """Report should not crash if there are zero matched pairs."""
        with tempfile.TemporaryDirectory() as tmp_dir:
            output_path = Path(tmp_dir) / "empty_match.xlsx"
            generate_excel_report(
                recon=self.recon,
                matched_rows=[],
                exception_rows=self.exception_rows,
                all_bank_txns=self.all_bank_txns,
                output_path=output_path,
            )
            assert output_path.exists()

    def test_report_works_with_empty_exception_rows(self):
        """Report should not crash if there are zero exceptions."""
        with tempfile.TemporaryDirectory() as tmp_dir:
            output_path = Path(tmp_dir) / "no_exceptions.xlsx"
            generate_excel_report(
                recon=self.recon,
                matched_rows=self.matched_rows,
                exception_rows=[],
                all_bank_txns=self.all_bank_txns,
                output_path=output_path,
            )
            assert output_path.exists()


# ===========================================================================
# PDF report — file output checks
# ===========================================================================

class TestPdfReport(ReportTestBase):

    def test_file_is_created(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            output_path = Path(tmp_dir) / "test_recon.pdf"
            generate_pdf_report(
                recon=self.recon,
                matched_rows=self.matched_rows,
                exception_rows=self.exception_rows,
                output_path=output_path,
            )
            assert output_path.exists()

    def test_file_is_non_empty(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            output_path = Path(tmp_dir) / "test_recon.pdf"
            generate_pdf_report(
                recon=self.recon,
                matched_rows=self.matched_rows,
                exception_rows=self.exception_rows,
                output_path=output_path,
            )
            assert output_path.stat().st_size > 0

    def test_file_starts_with_pdf_magic_bytes(self):
        """A valid PDF file always starts with the bytes b'%PDF'."""
        with tempfile.TemporaryDirectory() as tmp_dir:
            output_path = Path(tmp_dir) / "test_recon.pdf"
            generate_pdf_report(
                recon=self.recon,
                matched_rows=self.matched_rows,
                exception_rows=self.exception_rows,
                output_path=output_path,
            )
            first_bytes = output_path.read_bytes()[:4]
            assert first_bytes == b"%PDF"

    def test_returns_output_path(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            output_path = Path(tmp_dir) / "test_recon.pdf"
            returned = generate_pdf_report(
                recon=self.recon,
                matched_rows=self.matched_rows,
                exception_rows=self.exception_rows,
                output_path=output_path,
            )
            assert returned == output_path

    def test_output_path_parent_created_if_missing(self):
        """generate_pdf_report should create any missing parent directories."""
        with tempfile.TemporaryDirectory() as tmp_dir:
            nested_path = Path(tmp_dir) / "new_folder" / "report.pdf"
            generate_pdf_report(
                recon=self.recon,
                matched_rows=self.matched_rows,
                exception_rows=self.exception_rows,
                output_path=nested_path,
            )
            assert nested_path.exists()

    def test_report_works_with_no_exceptions(self):
        """PDF should build without errors when there are no exceptions."""
        with tempfile.TemporaryDirectory() as tmp_dir:
            output_path = Path(tmp_dir) / "no_exc.pdf"
            generate_pdf_report(
                recon=self.recon,
                matched_rows=self.matched_rows,
                exception_rows=[],
                output_path=output_path,
            )
            assert output_path.exists()

    def test_report_works_with_no_matched_rows(self):
        """PDF should build without errors when there are no matched rows."""
        with tempfile.TemporaryDirectory() as tmp_dir:
            output_path = Path(tmp_dir) / "no_match.pdf"
            generate_pdf_report(
                recon=self.recon,
                matched_rows=[],
                exception_rows=self.exception_rows,
                output_path=output_path,
            )
            assert output_path.exists()

    def test_pdf_truncates_exception_detail_at_20_rows(self):
        """PDF generator must not crash when there are more than 20 exceptions."""
        # Create 25 fake exception rows — over the MAX_EXCEPTION_DETAIL_ROWS limit
        many_exceptions = [
            {
                "source": "BANK",
                "exception_type": "BANK_ONLY",
                "txn_date": date(2026, 3, idx + 1),
                "description": f"Unmatched transaction {idx + 1}",
                "amount": 100.0 * (idx + 1),
                "reference": None,
            }
            for idx in range(25)
        ]
        with tempfile.TemporaryDirectory() as tmp_dir:
            output_path = Path(tmp_dir) / "many_exc.pdf"
            generate_pdf_report(
                recon=self.recon,
                matched_rows=self.matched_rows,
                exception_rows=many_exceptions,
                output_path=output_path,
            )
            assert output_path.exists()
            assert output_path.stat().st_size > 0