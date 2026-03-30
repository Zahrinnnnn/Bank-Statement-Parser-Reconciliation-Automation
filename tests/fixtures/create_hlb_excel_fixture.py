"""
create_hlb_excel_fixture.py — Generates a sample HLB Excel statement fixture.

Run once:
    python tests/fixtures/create_hlb_excel_fixture.py
"""

from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

OUTPUT_PATH = Path(__file__).parent / "sample_hlb.xlsx"


def create_hlb_excel() -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Statement"

    # Metadata rows (like a real HLB export)
    ws.append(["HONG LEONG BANK BERHAD"])
    ws.append(["Account Number:", "1234567890"])
    ws.append(["Account Holder:", "ZAHRIN BIN JASNI"])
    ws.append(["Statement Period:", "01/03/2026 - 31/03/2026"])
    ws.append([])  # blank row

    # Header row
    headers = ["Date", "Transaction Description", "Withdrawal (DR)", "Deposit (CR)", "Balance"]
    ws.append(headers)

    # Style the header row
    header_row_number = ws.max_row
    header_fill = PatternFill("solid", fgColor="003399")
    for cell in ws[header_row_number]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Transaction rows
    transactions = [
        ("03/03/2026", "FPX SHOPEE PAYMENT",           "250.00",  "",        "4750.00"),
        ("05/03/2026", "SALARY CREDIT ACME SDN BHD",   "",        "3500.00", "8250.00"),
        ("08/03/2026", "IBG TRANSFER OUT AHMAD ALI",   "1000.00", "",        "7250.00"),
        ("15/03/2026", "IBFT IN FROM SITI HASSAN",     "",        "500.00",  "7750.00"),
        ("18/03/2026", "TT789012 UTILITY TNB PAYMENT", "156.30",  "",        "7593.70"),
        ("20/03/2026", "CHEQ000456 RENT PAYMENT",      "1500.00", "",        "6093.70"),
        ("25/03/2026", "BONUS CREDIT ACME SDN BHD",    "",        "800.00",  "6893.70"),
        ("31/03/2026", "SERVICE CHARGE",               "10.00",   "",        "6883.70"),
    ]

    for row in transactions:
        ws.append(list(row))

    # Set column widths
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 14

    wb.save(str(OUTPUT_PATH))
    print(f"Created: {OUTPUT_PATH}")


if __name__ == "__main__":
    create_hlb_excel()
